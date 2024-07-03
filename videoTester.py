import os
import cv2
import numpy as np
import faceRecognition as fr
import smtplib
from email.message import EmailMessage
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime

# Load the trained face recognizer
face_recognizer = cv2.face.LBPHFaceRecognizer_create()
face_recognizer.read("C:/Users/Soundara Barath/Downloads/FaceRecognition-master/FaceRecognition-master/trainingData.yml")

# Define names corresponding to the labels recognized by the model
names = {0: "Riswanth", 1: "Dharneesh", 2: "Soundar", 3: "Mukul",4:"Rajesh"}

# Open the Excel workbook or create a new one if it doesn't exist
if os.path.exists('history.xlsx'):
    wb = openpyxl.load_workbook('history.xlsx')
    sheet1 = wb.active
else:
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = 'Sheet1'
    sheet1.append(['Name', 'Date', 'Time'])
    # Apply style for unknown names
    unknown_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for cell in sheet1['A1:C1']:
        for c in cell:
            c.fill = unknown_fill
            c.font = Font(bold=True)

row = sheet1.max_row + 1  # Start writing data from next row

# Initialize the video capture
cap = cv2.VideoCapture(0)

while True:
    # Read a frame from the video capture
    ret, test_img = cap.read()

    # Detect faces in the captured frame
    faces_detected, gray_img = fr.faceDetection(test_img)

    # Iterate through detected faces
    for (x, y, w, h) in faces_detected:
        # Draw rectangle around the face
        cv2.rectangle(test_img, (x, y), (x+w, y+h), (255, 0, 0), thickness=7)

        # Extract region of interest (ROI) from the grayscale image
        roi_gray = gray_img[y:y+h, x:x+w]

        # Predict the label and confidence for the face region
        label, confidence = face_recognizer.predict(roi_gray)
        print("confidence:", confidence)

        # Draw rectangle around the detected face
        fr.draw_rect(test_img, (x, y, w, h))

        # If confidence is above a threshold, mark as "Unknown" and send email
        if confidence > 50:
            predicted_name = "Unknown"
            label = "Unknown"

            # Sending an email for unknown face
            Sender_Email = "vmsoundarabharath@gmail.com"
            Recipient_Email = "vmsoundarabharath2003@gmail.com"
            Password = "oylc wozq lqzx jdkw"

            newMessage = EmailMessage()
            newMessage['Subject'] = "Alert: Unknown person detected"
            newMessage['From'] = Sender_Email
            newMessage['To'] = Recipient_Email

            # Attaching the image to the email
            with open('opencv.png', 'rb') as f:
                image_data = f.read()
                image_name = f.name  # File name without path
            newMessage.add_attachment(image_data, maintype='image', subtype='jpeg', filename=image_name)

            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                smtp.login(Sender_Email, Password)
                smtp.send_message(newMessage)

            print("Email sent successfully!")
        else:
            predicted_name = names.get(label, "Unknown")

        print("label:", label)
        fr.put_text(test_img, predicted_name, x, y)
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")
        
        # Write data to Excel and apply style for unknown names
        if predicted_name == "Unknown":
            unknown_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for cell in sheet1[f'A{row}:C{row}']:
                for c in cell:
                    c.fill = unknown_fill
            sheet1.append([predicted_name, datetime.today().date().strftime("%Y-%m-%d"), current_time])
        else:
            sheet1.append([predicted_name, datetime.today().date().strftime("%Y-%m-%d"), current_time])
        row += 1

    # Resize the image for display
    resized_img = cv2.resize(test_img, (640, 480))
    # Display the image
    cv2.imshow('Face Recognition', resized_img)

    # Break the loop if 'q' key is pressed
    if cv2.waitKey(10) == ord('q'):
        break

# Release video capture and close OpenCV windows
cap.release()
cv2.destroyAllWindows()

# Save the workbook
wb.save('history.xlsx')
